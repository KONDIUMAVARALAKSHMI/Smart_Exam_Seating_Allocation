from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from .forms import *
from django.utils import timezone
from django.contrib import messages
from .models import AddStudent, AddexamHall, Examallotment, AddFaculty, AdminAnnounce,AddTimeTable
from django.conf import settings
from django.core.mail import send_mail
from django.shortcuts import get_list_or_404
from itertools import chain
import secrets
import string
import random
from django.db.models import Q
import csv
from django.http import HttpResponse
from django.db import IntegrityError
import pandas as pd
from .forms import ExcelUploadForm
from django.contrib import messages
from .pdffile import *
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
import re
from reportlab.lib.styles import ParagraphStyle
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from io import BytesIO
from .models import AddStudent

from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak, KeepTogether
)
from reportlab.platypus import KeepTogether
# adminapp/views.py  (or wherever build_attendance_table lives)
import os                       # ← add this line
from io import BytesIO
from django.conf import settings
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Image, Spacer, PageBreak
)

import pandas as pd
import re
import string
import secrets
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import AddStudent
from .forms import ExcelUploadForm

# … the rest of your imports


from reportlab.lib.styles import getSampleStyleSheet

# To Delete Tables Data in Database
#AddTimeTable.objects.all().delete()
#Examallotment.objects.all().delete()
# AddexamHall.objects.all().delete()
#AddStudent.objects.all().delete()
#AddFaculty.objects.all().delete()
#Room.objects.all().delete()




from django.shortcuts import redirect
from django.contrib import messages
from .models import AddStudent

def delete_all_students(request):
    AddStudent.objects.all().delete()
    messages.success(request, "All students have been deleted successfully.")
    return redirect('/viewstudents')  # direct URL instead of name
# replace with your student list page URL name


def addtimetable(request):
    if request.method == 'POST':
        form = AddTimeTableForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('viewtimetable')  
    else:
        form = AddTimeTableForm()
    return render(request, 'addtimetable.html', {'form': form})

def viewtimetable(request):
    timetable_entries = AddTimeTable.objects.all()
    return render(request, 'viewtimetable.html', {'timetable_entries': timetable_entries})





# templates
INDEXPAGE = "index.html"
ADMINLOGINPAGE = "adminlogin.html"
ADMINHOMEPAGE = "adminhome.html"
ADDSTUDENTSPAGE = "addstudents.html"
ADDEXAMHALLSPAGE = "addexamhalls.html"
VIEWSTUDENTSPAGE = "viewstudents.html"
ADDFACULTYPAGE = "addfaculty.html"
VIEWFACULTYPAGE = "viewfaculty.html"
ADDANNOUNCEMENTPAGE = "addannouncement.html"
ADDTIMETABLEPAGE = "addtimetable.html"
VIEWTIMEPABLEPAGE = "viewtimetable.html"
# Create your views here.


def index(req):
    return render(req, INDEXPAGE)


def adminlogin(req):
    context = {}
    context['form'] = AdminlogForm()
    if req.method == "POST":
        form = AdminlogForm(req.POST)
        if form.is_valid():
            adminemail = form.cleaned_data['adminemail']
            adminpassword = form.cleaned_data['adminpassword']
            if adminemail == "admin@gmail.com" and adminpassword == "admin":
                req.session['adminemail'] = adminemail
                return render(req, ADMINHOMEPAGE)
            else:
                messages.warning(req, "Admin Credentials are not Valid......!")
                return render(req, ADMINLOGINPAGE, context)
    return render(req, ADMINLOGINPAGE, context)



import pandas as pd
import re
import string
import secrets
from django.contrib import messages
from django.shortcuts import redirect, render

def addstudents(req):
    context = {}
    if req.method == "POST":
        form = ExcelUploadForm(req.POST, req.FILES)
        if form.is_valid():
            excel_file = req.FILES['excel_file']
            if excel_file.name.endswith(('.xls', '.xlsx')):
                try:
                    # Read with correct header row (7th row in Excel -> index 6)
                    df = pd.read_excel(excel_file, header=6)
                    df.columns = df.columns.str.strip()  # Clean header names

                    # Confirm required columns exist
                    required_cols = ['Roll No', 'Name', 'department', 'Email', 'Student Phone']
                    for col in required_cols:
                        if col not in df.columns:
                            messages.error(req, f"Missing column: {col}")
                            return redirect('addstudents')

                    new_students = 0
                    for index, row in df.iterrows():
                        rollnumber = str(row['Roll No']).strip()
                        name = row['Name']
                        department = 'CSE'  # Default value, ignore Excel
                        email = row['Email']
                        contact = re.sub(r'\D', '', str(row['Student Phone']))
                        year = str(row['Year']).strip()
                        semester = str(row['Semester']).strip()       # Default value

                        if AddStudent.objects.filter(rollnumber=rollnumber).exists():
                            continue

                        random_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))

                        AddStudent.objects.create(
                            rollnumber=rollnumber,
                            name=name,
                            department=department,
                            email=email,
                            contact=contact,
                            year=year,
                            semester=semester,
                            password=random_password
                        )
                        new_students += 1

                    if new_students > 0:
                        messages.success(req, f"{new_students} students added successfully from Excel file.")
                    else:
                        messages.info(req, "No new students were added. All roll numbers already exist.")

                    return redirect('addstudents')

                except Exception as e:
                    messages.error(req, f"Error processing Excel file: {e}")
            else:
                messages.error(req, "Please upload a valid .xls or .xlsx Excel file.")
    else:
        form = ExcelUploadForm()

    context['form'] = form
    return render(req, 'addstudents.html', context)




def addexamhalls(request):
    if request.method == "POST":
        form = AddexamhallForm(request.POST)
        if form.is_valid():
            # ── pull cleaned data ───────────────────────────────
            date        = form.cleaned_data['Date']
            subject1    = form.cleaned_data['subject1']
            subject2    = form.cleaned_data['subject2']
            noofrooms   = form.cleaned_data['noofrooms']
            noofbenches = form.cleaned_data['noofbenches']
            students_per_bench = int(form.cleaned_data['students_per_bench'])

            selected_room_ids = request.POST.getlist('rooms')
            rooms = Room.objects.filter(id__in=selected_room_ids)
            selected_rooms = [room.room_number for room in rooms]

            # sanity-check: rooms chosen = number entered
            if len(selected_rooms) != noofrooms:
                form.add_error('rooms', f"Please select exactly {noofrooms} rooms.")
            else:
                total_benches = noofrooms * noofbenches
                total_seats_available = total_benches * students_per_bench
                total_students = AddStudent.objects.count()

                if total_seats_available < total_students:
                    messages.error(
                        request,
                        "Not enough seats available for the total number of students."
                    )
                else:
                    # ✅ DELETE previous exam hall records
                    AddexamHall.objects.all().delete()

                    # ✅ Create new exam hall
                    AddexamHall.objects.create(
                        date               = date,
                        subject1           = subject1,
                        subject2           = subject2,
                        noofrooms          = noofrooms,
                        noofbenches        = noofbenches,
                        total_benches      = total_benches,
                        total_seats        = total_seats_available,
                        rooms_list         = ",".join(selected_rooms),
                        students_per_bench = students_per_bench,
                    )

                    messages.success(request, "Exam hall added successfully.")
                    return redirect('addexamhalls')
        # else: form not valid → fall through to re-render with errors
    else:
        form = AddexamhallForm()

    rooms = Room.objects.all()
    return render(request, 'addexamhalls.html', {'form': form, 'rooms': rooms})


def delete(req, id):
    print(id)
    AddStudent.objects.filter(id=id).delete()
    return redirect("viewstudents")


def deletefaculty(req,id):
    AddFaculty.objects.filter(id=id).delete()
    return redirect("viewfaculty")



from django.shortcuts import redirect
from django.db import transaction

def setseatallotment(request):
    BENCHES_PER_ROOM = 18  # total benches
    SEATS_PER_BENCH = 3    # [Sem7, Sem5, Sem7]
    STUDENTS_PER_ROOM = BENCHES_PER_ROOM * SEATS_PER_BENCH  # 54

    # fetch students
    sem7_students = list(
        AddStudent.objects.filter(semester__in=['7', 7]).order_by("rollnumber", "id")
    )
    sem5_students = list(
        AddStudent.objects.filter(semester__in=['5', 5]).order_by("rollnumber", "id")
    )

    @transaction.atomic
    def _run():
        # clear previous
        Examallotment.objects.all().delete()

        halls = list(AddexamHall.objects.all())

        sem5_idx, sem7_idx = 0, 0

        for hall in halls:
            room_numbers_raw = (hall.rooms_list or "").split(',')
            room_numbers = [r.strip() for r in room_numbers_raw if r.strip()]
            if not room_numbers:
                continue

            if hall.noofrooms and hall.noofrooms > 0:
                room_numbers = room_numbers[:hall.noofrooms]

            for room_count, room_num in enumerate(room_numbers, start=1):
                room_display = f"Room{room_num}"

                # Decide distribution
                if room_count % 2 == 1:  # odd → 30 sem5, 24 sem7
                    sem5_quota, sem7_quota = 30, 24
                else:                   # even → 30 sem7, 24 sem5
                    sem5_quota, sem7_quota = 24, 30

                # pick required students
                room_sem5 = sem5_students[sem5_idx: sem5_idx + sem5_quota]
                sem5_idx += sem5_quota
                room_sem7 = sem7_students[sem7_idx: sem7_idx + sem7_quota]
                sem7_idx += sem7_quota

                # arrange benches
                benches = []
                s5_ptr, s7_ptr = 0, 0

                for bench_no in range(1, BENCHES_PER_ROOM + 1):
                    # default pattern [sem7, sem5, sem7]
                    left, middle, right = None, None, None

                    # fill left seat
                    if s7_ptr < len(room_sem7):
                        left = room_sem7[s7_ptr]
                        s7_ptr += 1

                    # fill middle seat
                    if s5_ptr < len(room_sem5):
                        middle = room_sem5[s5_ptr]
                        s5_ptr += 1

                    # fill right seat
                    if s7_ptr < len(room_sem7):
                        right = room_sem7[s7_ptr]
                        s7_ptr += 1

                    benches.append((left, middle, right))

                # save into DB
                for bno, (s1, s2, s3) in enumerate(benches, start=1):
                    def add_allot(stu, seat_no):
                        if stu:
                            Examallotment.objects.create(
                                Student_Id=stu.rollnumber,
                                department=stu.department or "CSE",
                                RoomNo=room_display,
                                BenchNo=f"Bench{bno}",
                                SeatNumber=f"Seat{seat_no}",
                                date=hall.date,
                                starttime=getattr(hall, 'starttime', None),
                                endtime=getattr(hall, 'endtime', None),
                            )

                    add_allot(s1, 1)  # left (Sem7)
                    add_allot(s2, 2)  # middle (Sem5)
                    add_allot(s3, 3)  # right (Sem7)

    _run()
    return redirect("viewallotedstudents")







def viewallotedstudents(request):
    Exam_alloted_student = Examallotment.objects.all()
    return render(request, "viewallotedstudents.html", {'Exam_alloted_student': Exam_alloted_student})

def viewstudents(req):
    all_students = AddStudent.objects.all()
    return render(req, VIEWSTUDENTSPAGE, {'all_students': all_students})


def addfaculty(req):
    context = {}
    context['form'] = AddFacultyForm()
    if req.method == "POST":
        form = AddFacultyForm(req.POST)
        if form.is_valid():
            length = 8
            characters = string.ascii_letters + string.digits

            # Generate a random password
            random_password = ''.join(secrets.choice(characters) for _ in range(length))
            print("Random Password:", random_password)

            # Extracting form data
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            contact = form.cleaned_data['contact']
            branch = form.cleaned_data['branch']
            subject = form.cleaned_data['subject']
            semester = form.cleaned_data['semester']
            year = form.cleaned_data['year']
            #image = form.cleaned_data['image']
            #profilename = image.name

            try:
                # Attempt to save faculty member
                faculty = AddFaculty.objects.create(
                    name=name,
                    email=email,
                    contact=contact,
                    branch=branch,
                    subject=subject,
                    semester=semester,
                    year=year,
                    #image=image,
                    #profilename=profilename,
                    password=random_password
                )
            # # Mail Code
            #     # Sending email with login credentials
            #     subject = "Exam Details"
            #     cont = f'Dear {name}'
            #     KEY = f' Branch : {branch}\n'
            #     m1 = f"Your Login Credentials Username : {email}  & password {random_password}"
            #     m2 = "Thanking you"
            #     m3 = "Regards"
            #     m4 = "Admin."

            #     email_from = settings.EMAIL_HOST_USER
            #     recipient_list = [email]
            #     text = cont + '\n' + KEY + '\n' + m1 + '\n' + m2 + '\n' + m3 + '\n' + m4
            #     send_mail(subject, text, email_from, recipient_list, fail_silently=False)

                messages.success(req, "Faculty added successfully")
            except IntegrityError:
                # If email already exists, handle the exception
                messages.warning(req, "A faculty member with the same email already exists")

    return render(req, 'addfaculty.html', context)

#def addannouncement(req):
 #   two_days_content = timezone.now()-timezone.timedelta(days=2)
  #  messages_to_delete = AdminAnnounce.objects.filter(annuncementdate=two_days_content)
  #  all_messages = AdminAnnounce.objects.all()
  #  context = {}
   # context['form'] = AdminAnnouncement()
#
 #      form = AdminAnnouncement(req.POST)
  #      print(form.is_valid())
   #     if form.is_valid():
    #       adminemail = req.session['adminemail']
     #       data = AdminAnnounce(
     #           announcement=announcement,
     #           senderemail=adminemail
     #       )
     #       data.save()

     #       # Correct syntax for passing context to the template
     #       return render(req, ADDANNOUNCEMENTPAGE, {'form': AdminAnnouncement(), 'all_messages': all_messages})

   # return render(req, ADDANNOUNCEMENTPAGE, {'form': AdminAnnouncement(), 'all_messages': all_messages})



def viewfaculty(req):
    all_faculty = AddFaculty.objects.all()
    return render(req, VIEWFACULTYPAGE, {'all_faculties': all_faculty})



# ---------------- DOWNLOAD EXCEL ----------------
def download_details(req):
    details_data = Examallotment.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="details.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    # College Name
    worksheet.merge_cells('A1:I1')
    college_name_cell = worksheet['A1']
    college_name_cell.value = "Aditya University"
    college_name_cell.font = Font(bold=True, size=18)
    college_name_cell.alignment = Alignment(horizontal='center')

    # Venue
    worksheet.merge_cells('A2:I2')
    venue_cell = worksheet['A2']
    venue_cell.value = "Venue: BGB"
    venue_cell.font = Font(bold=True)
    venue_cell.alignment = Alignment(horizontal='center')

    # Headers
    header_row = [
        'Branch', 'RoomNo', 'BenchNo', 'SeatNumber',
        'Student_Id', 'Date', 'Start Time', 'End Time'
    ]
    worksheet.append(header_row)
    for cell in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row,
                                     min_col=1, max_col=len(header_row)):
        for c in cell:
            c.font = Font(bold=True)

    # Data rows
    for detail in details_data:
        data_row = [
            detail.department, detail.RoomNo, detail.BenchNo, detail.SeatNumber,
            detail.Student_Id, detail.date, detail.starttime, detail.endtime
        ]
        worksheet.append(data_row)

    workbook.save(response)
    return response


# ---------------- PDF ATTENDANCE ----------------
def attendance_sheet_home(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=36, rightMargin=36,
        topMargin=30, bottomMargin=20
    )
    elements = []
    styles = getSampleStyleSheet()

    # Department mapping from roll number
    dept_mapping = {
        "DS": "Data Science",
        "IT": "Information Technology",
        # Add more as needed
    }

    # Get all unique rooms
    all_rooms = (
        Examallotment.objects
        .values_list('RoomNo', flat=True)
        .distinct()
        .order_by('RoomNo')
    )

    first_section = True

    for room_label in all_rooms:
        numeric_room = re.sub(r'[^0-9]', '', str(room_label) or "")
        hall = AddexamHall.objects.filter(rooms_list__icontains=numeric_room).first()

        # Get all students in the room
        student_ids = (
            Examallotment.objects
            .filter(RoomNo__icontains=numeric_room)
            .values_list('Student_Id', flat=True)
            .distinct()
        )
        students = (
            AddStudent.objects
            .filter(rollnumber__in=student_ids)
            .order_by('rollnumber')
        )

        # Group students by department code from roll number
        dept_groups = {}
        for stu in students:
            dept_code = stu.rollnumber[5:7]  # adjust index if needed
            dept_name = dept_mapping.get(dept_code, dept_code)
            dept_groups.setdefault(dept_name, []).append(stu)

        # Build one attendance sheet per department for this room
        for dept_name, dept_students in dept_groups.items():
            elements += build_attendance_table(
                hall, dept_students, room_label, dept_name, styles, add_page_break=not first_section
            )
            first_section = False

    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type="application/pdf")


def build_attendance_table(hall, student_list, room, dept_name, styles, add_page_break=True):
    elements = []
    if add_page_break:
        elements.append(PageBreak())

    # Logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'adlogo.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=350, height=80)
        logo.hAlign = "CENTER"
        elements.append(logo)
        elements.append(Spacer(1, 6))

    semester = "III Sem"  # change if needed

    centered_heading = ParagraphStyle(
        name='CenteredHeading', parent=styles['Heading2'], alignment=1
    )
    elements.append(Paragraph(f"<b>Department of {dept_name}</b>", centered_heading))
    elements.append(Spacer(1, 6))

    # Metadata tables
    left_metadata = [
        [Paragraph("<b>Venue:</b> BGB", styles['BodyText'])],
        [Paragraph(f"<b>Room Number:</b> {room}", styles['BodyText'])]
    ]
    left_table = Table(left_metadata, hAlign='LEFT')

    date_str = hall.date.strftime('%d-%m-%Y') if hall and hall.date else "N/A"
    right_metadata = [
        [Paragraph(f"<b>Date:</b> {date_str}", styles['BodyText'])],
        [Paragraph(f"<b>Semester:</b> {semester}", styles['BodyText'])]
    ]
    right_table = Table(right_metadata, hAlign='RIGHT')

    combined = Table([[left_table, right_table]], colWidths=[350, 150])
    combined.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    elements.extend([combined, Spacer(1, 12)])

    # Title
    elements.append(Paragraph("<b>Attendance Sheet</b>", styles['Heading2']))
    elements.append(Spacer(1, 12))

    # Attendance table
    table_data = [["S.No", "Roll No", "Name", "Signature"]]
    if student_list:
        for idx, stu in enumerate(student_list, start=1):
            table_data.append([idx, stu.rollnumber, stu.name, ""])
    else:
        table_data.append(["-", "-", "No students assigned", ""])

    att_table = Table(table_data, colWidths=[50, 120, 250, 80], repeatRows=1)
    att_table.setStyle(TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('ALIGN', (0, 0), (1, -1), 'CENTER'),  # S.No & Roll No centered
    ('ALIGN', (2, 0), (2, -1), 'LEFT'),    # Name column left-aligned
    ('ALIGN', (3, 0), (3, -1), 'CENTER'),  # Signature column centered
    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ('TOPPADDING', (0, 0), (-1, -1), 6),
    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
]))

    elements.append(att_table)

    # Signature lines
    elements.append(Spacer(1, 36))
    signature_tbl = Table([
        ["", "", "__________________________", "__________________________"],
        ["", "", "Invigilator 1 Signature", "Invigilator 2 Signature"],
    ], colWidths=[50, 150, 150, 150])
    signature_tbl.setStyle(TableStyle([
        ('ALIGN', (2, 0), (3, 0), 'CENTER'),
        ('ALIGN', (2, 1), (3, 1), 'CENTER'),
        ('FONTSIZE', (2, 1), (3, 1), 10),
        ('TOPPADDING', (2, 0), (3, 0), 20),
    ]))
    elements.append(signature_tbl)

    return elements
